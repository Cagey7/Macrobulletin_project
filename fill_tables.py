from datetime import datetime
import os
import calendar
from openpyxl import Workbook
from taldau import Automation

class FillTable(Automation):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.excel_path = os.path.join(os.getcwd(), "bulletin", "macroeconomics", "static", "macroeconomics", "tables")


    def consumer_price_index_fill_table(self, start_month=None, start_year=None, end_month=None, end_year=None):
        excel_file_names = [["consumer_price_index1.xlsx", "отчетный период к соответствующему периоду прошлого года"], 
                            ["consumer_price_index2.xlsx", "отчетный период к предыдущему периоду"]]
        
        for name in excel_file_names:
            workbook = Workbook()
            sheet = workbook.active

            if end_month is None or end_year is None:
                current_date = datetime.now()
                end_month = current_date.month
                end_year = current_date.year
            
            if start_month is None or start_year is None:
                current_date = datetime.now()
                start_month = 12
                start_year = current_date.year - 2
            
            start_last_day = calendar.monthrange(start_year, start_month)[1]
            end_last_day = calendar.monthrange(end_year, end_month)[1]

            months = [
                ["янв", 'Январь'], 
                ["фев", 'Февраль'], 
                ["мар", 'Март'], 
                ["апр", 'Апрель'], 
                ["май", 'Май'], 
                ["июн", 'Июнь'], 
                ["июл", 'Июль'], 
                ["авг", 'Август'], 
                ["сен", 'Сентябрь'], 
                ["окт", 'Октябрь'], 
                ["ноя", 'Ноябрь'], 
                ["дек", 'Декабрь']
            ]

            bull_name = [
                ["Потребительские цены", 'Товары и услуги'], 
                ["Продовольственные товары", 'Продовольственные товары'], 
                ["Непродовольственные товары", 'Непродовольственные товары'], 
                ["Платные услуги", 'Платные услуги'], 
                ["Цены производителей", 'Промышленность']
            ]


            # Создание мест списку для поиска
            condition = ",".join(["%s"] * len(bull_name))


            excel_path = os.path.join(self.excel_path, name[0])

            join_views = f"""
            SELECT 
                activity_type, 
                value, 
                description, 
                created_at
            FROM 
                consumer_price_index_month 
            WHERE 
                region = 'РЕСПУБЛИКА КАЗАХСТАН' 
                AND periods_correlation = '{name[1]}'  
                AND activity_type IN ({condition}) 
                AND created_at BETWEEN '{start_year}-{start_month}-{start_last_day}' AND '{end_year}-{end_month}-{end_last_day}'
            UNION
            SELECT 
                activity_type, 
                value, 
                description, 
                created_at
            FROM 
                producer_price_index_month 
            WHERE 
                region = 'РЕСПУБЛИКА КАЗАХСТАН' 
                AND periods_correlation = '{name[1]}' 
                AND countries = 'Всего'
                AND activity_type = 'Промышленность'
                AND industrial_products_list = 'Всего'
                AND created_at BETWEEN '{start_year}-{start_month}-{start_last_day}' AND '{end_year}-{end_month}-{end_last_day}';
            """

            self.cur.execute(join_views, [item[1] for item in bull_name])
            sorted_data = self.cur.fetchall()
            dates = [item[3] for item in sorted_data]
            latest_month = max(dates).month
            latest_year = max(dates).year

            if latest_month < end_month:
                end_month = latest_month
            if latest_year < end_year:
                end_year = latest_year

            
            months_list = [""]

            data_for_excel = []
            for bull_name_i, name in enumerate(bull_name):
                index_data = []
                index_data.append(name[0])
                for year in range(start_year, end_year+1):
                    for i, m in enumerate(months, start=1):
                        if year == start_year and i < start_month: continue
                        if year == end_year and i > end_month: continue
                        # получешние месяцев для заголовка
                        if bull_name_i == 0:
                            months_list.append(m[0])
                        
                        data_not_exist = True
                        for data in sorted_data:
                            t_month, t_year, _ = data[2].split()
                            t_year = int(t_year)
                            if name[1] == data[0] and t_month == m[1] and t_year == year:
                                index_data.append(float(data[1]))
                                data_not_exist = False
                        if data_not_exist:
                            index_data.append("")
                data_for_excel.append(index_data)

            # заголовок
            year_list = list(range(start_year, end_year + 1))
            start_cell = 2
            end_cell = 12-start_month+start_cell
            sheet.cell(row=1, column=1, value="Показатели")
            for i, year in enumerate(year_list):
                sheet.cell(row=1, column=start_cell, value=year)
                if start_year == end_year and start_month == end_month: break
                sheet.merge_cells(start_row=1, start_column=start_cell, end_row=1, end_column=end_cell)
                start_cell = end_cell+1
                if i+2 == len(year_list):
                    end_cell += end_month
                else:
                    end_cell += 12



            sheet.append(months_list)
            for data in data_for_excel:
                sheet.append(data)
            workbook.save(excel_path)
            workbook.close()


    def avsalary_byactivity_fill_table(self, start_year=None, end_year=None):
        excel_path = os.path.join(self.excel_path, "average_salary_by_economic_activity_in_thousand_tenge.xlsx")
        workbook = Workbook()
        sheet = workbook.active

        if end_year is None:
            current_date = datetime.now()
            end_year = current_date.year
        
        if start_year is None:
            current_date = datetime.now()
            start_year = current_date.year - 3

        bull_name = [
            ["Республика Казахстан","Всего"],
            ["Сельское, лесное и рыбное хозяйство","Сельское, лесное и рыбное хозяйство"],
            # ["Промышленность", "Промышленность"],
            ["Горнодобывающая промышленность","Горнодобывающая промышленность и разработка карьеров"],
            ["Обрабатывающая промышленность","Обрабатывающая промышленность",],
            ["Снабжение электроэнергией","Снабжение электроэнергией, газом, паром, горячей водой и кондиционированным воздухом"],
            ["Водоснабжение","Водоснабжение; сбор, обработка и удаление отходов, деятельность по ликвидации загрязнений"],
            ["Строительство","Строительство"],
            ["Оптовая и розничная торговля","Оптовая и розничная торговля; ремонт автомобилей и мотоциклов"],
            ["Транспорт и складирование","Транспорт и складирование"],
            ["Предоставление услуг по проживанию и питанию","Предоставление услуг по проживанию и питанию"],
            ["Информация и связь","Информация и связь"],
            ["Финансовая и страховая деятельность","Финансовая и страховая деятельность"],
            ["Операции с недвижимым имуществом","Операции с недвижимым имуществом"],
            ["Профессиональная, научная и техническая деятельность","Профессиональная, научная и техническая деятельность"],
            ["Деятельность в области административного обслуживания","Деятельность в области административного и вспомогательного обслуживания"],
            ["Государственное управление и оборона","Государственное управление и оборона; обязательное  социальное обеспечение"],
            ["Образование","Образование"],
            ["Здравоохранение","Здравоохранение и социальное обслуживание населения"],
            ["Искусство, развлечения и отдых","Искусство, развлечения и отдых"],
            ["Предоставление прочих видов услуг","Предоставление прочих видов услуг"]
        ]


        # Создание мест списку для поиска
        condition = ",".join(["%s"] * len(bull_name))
    
        join_views = f'''
        SELECT 
            activity_type,
            value,
            description,
            created_at
        FROM 
            avmon_nom_wages_year 
        WHERE 
            region = 'РЕСПУБЛИКА КАЗАХСТАН' 
            AND activity_type IN ({condition}) 
            AND terrain_type = 'Всего' 
            AND enterprise_dimension = 'Всего' 
            AND gender = 'Всего' 
            AND created_at BETWEEN '{start_year}-01-01' AND '{start_year}-12-31'
        UNION
        SELECT 
            activity_type, 
            value, 
            description, 
            created_at 
        FROM
            avmon_nom_wages_quarter 
        WHERE 
            region = 'РЕСПУБЛИКА КАЗАХСТАН' AND activity_type IN ({condition})
            AND economic_sectors = 'Экономика в целом'
            AND created_at BETWEEN '{start_year+1}-01-01' AND '{end_year+1}-12-01';
        '''
        condition_values = []
        for _ in range(2):
            for item in bull_name:
                condition_values.append(item[1])

        self.cur.execute(join_views, condition_values)
        
        sorted_data = self.cur.fetchall()

        latest_date = max(sorted_data, key=lambda x: x[3])

        for data in sorted_data:
            if data == latest_date:
                last_quarter = int(data[2][0])
                break

        quarter_list = ["в тенге", "Год"]

        data_for_excel = []
        for bull_name_i, name in enumerate(bull_name):
            index_data = []
            index_data.append(name[0])
            data_not_exist = True
            for data in sorted_data:
                if len(data[2].split()) == 2:
                    if data[0] == name[1]:
                        index_data.append(float(round(data[1]/1000, 1)))
                        data_not_exist = False
            if data_not_exist:
                index_data.append("")
            
            for year in range(start_year+1, end_year+1):
                end_quarter = 5
                if year == end_year: end_quarter = last_quarter + 1
                for quarter in range(1, end_quarter):
                    data_not_exist = True
                    for data in sorted_data:
                        if data[2] == f"{quarter} квартал {year} г." and data[0] == name[1]:
                            index_data.append(float(round(data[1]/1000, 1)))
                            data_not_exist = False
                            if bull_name_i == 0:
                                quarter_list.append(f"{quarter} кв")
                    if data_not_exist:
                        index_data.append("")
            data_for_excel.append(index_data)

        # заголовок
        year_list = list(range(start_year+1, end_year+1))
        start_cell = 3
        end_cell = 6
        sheet.cell(row=1, column=1, value="")
        sheet.cell(row=1, column=2, value=start_year)
        for i, year in enumerate(year_list):
            sheet.cell(row=1, column=start_cell, value=year)
            sheet.merge_cells(start_row=1, start_column=start_cell, end_row=1, end_column=end_cell)
            start_cell = end_cell+1
            if i+2 == len(year_list):
                end_cell += last_quarter
            else:
                end_cell += 4
        
        sheet.append(quarter_list)
        for data in data_for_excel:
            sheet.append(data)
                
        workbook.save(excel_path)
        workbook.close()

fillTable = FillTable("testtest", "postgres", "123456")

#fillTable.avsalary_byactivity_fill_table()
#fillTable.consumer_price_index_fill_table()
fillTable.db_disconnect()