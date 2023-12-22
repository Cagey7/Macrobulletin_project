from datetime import datetime, timedelta
from datetime import date
from dateutil.relativedelta import relativedelta
import os
import calendar
from openpyxl import Workbook
from automation import Automation
from sql import *
from table_info import *



class FillTable(Automation):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        current_directory = os.path.abspath(os.getcwd())
        parent_directory = os.path.dirname(current_directory)
        self.excel_path = os.path.join(parent_directory, "bulletin", "macroeconomics", "static", "macroeconomics", "tables")

    def convert_number(self, data_type, num):
        if data_type == "tenge" or data_type == "thousand_tenge":
            return float(round(num/1000, 1))
        elif data_type == "percent":
            return float(round(num, 1))
        elif data_type == "million_tenge":
            rev_num = str(round(int(num/1000000)))[::-1]
            return ' '.join([rev_num[i:i+3] for i in range(0, len(rev_num), 3)])[::-1]
        elif data_type == "thousand_tenge_exp":
            if num < 100:
                return float(num)
            else:
                return float(round(num/1000, 1))

    def period_counter(self, data, period_name):
        for row in data:
            if period_name == "quarter":
                desc_parts = row[2].split()
                period = desc_parts[0] + " " + desc_parts[1]
                for key in quarter_list:
                    if period == key:
                        quarter_list[key] += 1
            elif period_name == "month_accum":
                desc_parts = row[2].split()
                period = desc_parts[0] + " " + desc_parts[1] + " " + desc_parts[2]
                for key in month_accum_list:
                    if period == key:
                        month_accum_list[key] += 1
        if period_name == "quarter":
            if quarter_list["1 квартал"] == 0:
                return 0
            period_parts = max(quarter_list, key=quarter_list.get).split()
            return period_parts[0]
        elif period_name == "month_accum":
            if month_accum_list["Январь - Январь"] == 0 and month_accum_list["Январь - Март"] == 0:
                return 0
            return max(month_accum_list, key=month_accum_list.get)

    def create_labor_stats(self, names, sql, table_name, index_name, start_year=None):
        excel_path = os.path.join(self.excel_path, table_name)

        if start_year is None:
            current_date = datetime.now()
            start_year = current_date.year - 4
        workbook = Workbook()
        sheet = workbook.active
        end_year = start_year + 3

        date_start_year = datetime(start_year, 1, 1)
        date_end_year = datetime(start_year + 4, 1, 1) - timedelta(days=1)
        date_start_quarter = datetime(start_year + 4, 1, 1)
        date_end_quarter = datetime.now()

        condition_values = []
        for name in names[2:21]:
            condition_values.append(name[1])

        self.cur.execute(
            sql, (
                date_start_year, date_end_year,
                date_start_year, date_end_year,
                tuple(condition_values),
                date_start_year, date_end_year,
                date_start_year, date_end_year,
                date_start_year, date_end_year,
                date_start_year, date_end_year,
                date_start_year, date_end_year,
                date_start_year, date_end_year,
                date_start_quarter, date_end_quarter,
                date_start_quarter, date_end_quarter,
                tuple(condition_values),
                date_start_quarter, date_end_quarter,
                date_start_quarter, date_end_quarter,
                date_start_quarter, date_end_quarter,
                date_start_quarter, date_end_quarter,
                date_start_quarter, date_end_quarter,                
            )
        )

        sorted_data = self.cur.fetchall()

        last_year_data = []
        for data in sorted_data:
            if len(data[2].split()) > 2:
                last_year_data.append(data)

        quarter = self.period_counter(last_year_data, "quarter")
        # quarter = 1

        data_for_excel = []
        for name in names:
            index_data = []
            index_data.append(name[0])
            data_not_exist = True

            for year in range(start_year, end_year+1):
                data_not_exist = True
                for data in sorted_data:
                    if data[2] == f"{year} год" and data[0] == name[1]:
                        if data[1] < 100:
                            index_data.append(float(data[1]))
                        else:
                            index_data.append(float(round(data[1]/1000, 1)))
                        data_not_exist = False
                if data_not_exist:
                    index_data.append("")

            if quarter != 0:
                data_not_exist = True
                for data in sorted_data:
                    if data[2] == f"{quarter} квартал {end_year+1} г." and data[0] == name[1]:
                        if data[1] < 100:
                            index_data.append(float(data[1]))
                        else:
                            index_data.append(float(round(data[1]/1000, 1)))
                        data_not_exist = False
                if data_not_exist:
                    index_data.append("")

            data_for_excel.append(index_data)

        #заголовок
        if quarter == 0:
            end_year_header = end_year+1
        else:
            end_year_header = end_year+2
        header1 = ["Показатели"]
        for year in range(start_year, end_year_header):
            if year == end_year+1:
                header1.append(str(year) + f"\n{quarter} кв.")
            else:
                header1.append(year)


        sheet.append(header1)
        for data in data_for_excel:
            sheet.append(data)

        workbook.save(excel_path)
        workbook.close()

        print(f"Таблица {index_name} создана")

    def create_quarter_table(self, names, sql, table_name, data_type, index_name, start_year=None):
        description = {"tenge": ["в тенге", "Год"], "percent": ["в %", "Год"], "thousand_tenge": ["в тыс тенге", "Год"]}
        excel_path = os.path.join(self.excel_path, table_name)
        workbook = Workbook()
        sheet = workbook.active

        date_end = datetime.now()
        end_year = date_end.year
        
        if start_year is None:
            start_year = date_end.year - 3
        
        date_start = date(int(start_year), 1, 1)

        condition_values = []

        for name in names:
            condition_values.append(name[1])

        self.cur.execute(
            sql, (
                tuple(condition_values),
                date_start, 
                date_start+timedelta(days=366),
                tuple(condition_values), 
                date_start+timedelta(days=367), 
                date_end
            )
        )
        
        sorted_data = self.cur.fetchall()

        latest_date = max(sorted_data, key=lambda x: x[3])
        for data in sorted_data:
            if data == latest_date:
                last_quarter = int(data[2][0])
                last_year = int(data[2].split()[-2])
                break
        

        data_for_excel = []
        for name_i, name in enumerate(names):
            index_data = []
            index_data.append(name[0])
            data_not_exist = True
            for data in sorted_data:
                if len(data[2].split()) == 2:
                    if data[0] == name[1]:
                        value = self.convert_number(data_type, data[1])
                        index_data.append(value)
                        data_not_exist = False
            if data_not_exist:
                index_data.append("")
            
            for year in range(start_year+1, last_year+1):
                end_quarter = 5
                if year == last_year: end_quarter = last_quarter + 1
                for quarter in range(1, end_quarter):
                    if name_i == 0:
                        description[data_type].append(f"{quarter} кв")
                    data_not_exist = True
                    for data in sorted_data:
                        if data[2] == f"{quarter} квартал {year} г." and data[0] == name[1]:
                            value = self.convert_number(data_type, data[1])
                            index_data.append(value)
                            data_not_exist = False
                    if data_not_exist:
                        index_data.append("")
            data_for_excel.append(index_data)
        
        year_list = list(range(start_year+1, last_year+1))
        start_cell = 3
        end_cell = 6
        sheet.cell(row=1, column=1, value="")
        sheet.cell(row=1, column=2, value=start_year)
        for i, year in enumerate(year_list):
            sheet.cell(row=1, column=start_cell, value=year)
            sheet.merge_cells(start_row=1, start_column=start_cell, end_row=1, end_column=end_cell)
            start_cell = end_cell+1
            if i+2 == len(year_list):
                end_cell += last_quarter
            else:
                end_cell += 4

        sheet.append(description[data_type])
        for data in data_for_excel:
            sheet.append(data)
                
        workbook.save(excel_path)
        workbook.close()

        print(f"Таблица {index_name} создана")

    def create_year_table(self, names, sql, table_name, data_type, index_name, start_year=None):
        excel_path = os.path.join(self.excel_path, table_name)

        if start_year is None:
            current_date = datetime.now()
            start_year = current_date.year - 4
        workbook = Workbook()
        sheet = workbook.active
        end_year = start_year + 3

        date_start_year = datetime(start_year, 1, 1)
        date_end_year = datetime(start_year + 4, 1, 1) - timedelta(days=1)
        date_start_quarter = datetime(start_year + 4, 1, 1)
        date_end_quarter = datetime.now()


        self.cur.execute(
            sql, (
                tuple([item[1] for item in names]),
                date_start_year, date_end_year,
                tuple([item[1] for item in names]),
                date_start_quarter, date_end_quarter,
            )
        )
        sorted_data = self.cur.fetchall()

        date_list = []
        for data in sorted_data:
            if data[3].year == end_year+1:
                date_list.append(data[3])
        max_date = max(date_list)
        
        for data in sorted_data:
            if data[3] == max_date:
                accum_quarter = data[2]
                break

        try:
            int(accum_quarter[:1])
            accum_quarter = "no accum quarter"
        except:
            pass
        

        data_for_excel = []
        for name in names:
            index_data = []
            index_data.append(name[0])
            data_not_exist = True

            for year in range(start_year, end_year+1):
                data_not_exist = True
                for data in sorted_data:
                    if data[2] == f"{year} год" and data[0] == name[1]:
                        value = self.convert_number(data_type, data[1])
                        index_data.append(value)
                        data_not_exist = False
                if data_not_exist:
                    index_data.append("")

            if accum_quarter != "no accum quarter":
                data_not_exist = True
                for data in sorted_data:
                    if data[2] == accum_quarter and data[0] == name[1]:
                        value = self.convert_number(data_type, data[1])
                        index_data.append(value)
                        data_not_exist = False
                if data_not_exist:
                    index_data.append("")
            
            data_for_excel.append(index_data)

        if data_type == "thousand_tenge":
            header1 = ["Показатели \n(тыс. тенге)"]
        elif data_type == "million_tenge":
            header1 = ["Показатели \n(млн тенге)"]

        #заголовок
        if accum_quarter == "no accum quarter":
            end_year_header = end_year+1
            for year in range(start_year, end_year_header):
                header1.append(year)
        else:

            end_year_header = end_year+2
            words = accum_quarter.split()
            if len(words) >= 3:
                acc_quarter_months = ' '.join(words[:3])
            for year in range(start_year, end_year_header):
                if year == end_year+1:
                    header1.append(str(year) + f"\n {acc_quarter_months}")
                else:
                    header1.append(year)


        sheet.append(header1)
        for data in data_for_excel:
            sheet.append(data)
                
        workbook.save(excel_path)
        workbook.close()

        print(f"Таблица {index_name} создана")

    def create_month_table(self, names, months, sql, table_name, reporting_period, index_name, start_year=None, start_month=None):
        excel_path = os.path.join(self.excel_path, table_name)

        workbook = Workbook()
        sheet = workbook.active

        current_date = datetime.now()
        end_month = current_date.month
        end_year = current_date.year
        
        if start_month is None or start_year is None:
            current_date = datetime.now()
            start_month = 12
            start_year = current_date.year - 2
        
        start_last_day = calendar.monthrange(start_year, start_month)[1]
        end_last_day = calendar.monthrange(end_year, end_month)[1]


        self.cur.execute(
            sql, (
                reporting_period,
                tuple([item[1] for item in names]),
                datetime(start_year, start_month, start_last_day), 
                datetime(end_year, end_month, end_last_day),
                reporting_period, 
                datetime(start_year, start_month, start_last_day), 
                datetime(end_year, end_month, end_last_day)
            )
        )
        sorted_data = self.cur.fetchall()


        dates = [item[3] for item in sorted_data]
        latest_month = max(dates).month
        latest_year = max(dates).year

        if latest_month < end_month:
            end_month = latest_month
        if latest_year < end_year:
            end_year = latest_year

        
        months_list = [""]

        data_for_excel = []
        for bull_name_i, name in enumerate(names):
            index_data = []
            index_data.append(name[0])
            for year in range(start_year, end_year+1):
                for i, m in enumerate(months, start=1):
                    if year == start_year and i < start_month: continue
                    if year == end_year and i > end_month: continue
                    # получешние месяцев для заголовка
                    if bull_name_i == 0:
                        months_list.append(m[0])
                    
                    data_not_exist = True
                    for data in sorted_data:
                        t_month, t_year, _ = data[2].split()
                        t_year = int(t_year)
                        if name[1] == data[0] and t_month == m[1] and t_year == year:
                            index_data.append(float(data[1]))
                            data_not_exist = False
                    if data_not_exist:
                        index_data.append("")
            data_for_excel.append(index_data)

        # заголовок
        year_list = list(range(start_year, end_year + 1))
        start_cell = 2
        end_cell = 12-start_month+start_cell
        sheet.cell(row=1, column=1, value="Показатели")
        for i, year in enumerate(year_list):
            sheet.cell(row=1, column=start_cell, value=year)
            if start_year == end_year and start_month == end_month: break
            sheet.merge_cells(start_row=1, start_column=start_cell, end_row=1, end_column=end_cell)
            start_cell = end_cell+1
            if i+2 == len(year_list):
                end_cell += end_month
            else:
                end_cell += 12


        sheet.append(months_list)
        for data in data_for_excel:
            sheet.append(data)
        workbook.save(excel_path)
        workbook.close()

        print(f"Таблица {index_name} создана")


