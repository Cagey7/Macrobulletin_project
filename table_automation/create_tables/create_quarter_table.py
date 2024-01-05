import os
from openpyxl import Workbook
from filltable import FillTable
from datetime import datetime, timedelta

class CreateQuarterTable(FillTable):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


    def get_data_create_quarter_table(self, sql, names, minus_num):
        end_year = datetime.now().year
        
        start_year = end_year - minus_num
    
        date_start_year = datetime(start_year, 1, 1)
        date_end_year = datetime(start_year + 1, 1, 1) - timedelta(days=1)
        date_start_quarter = datetime(start_year + 1, 1, 1)
        date_end_quarter = datetime.now()

        condition_values = []

        for name in names:
            condition_values.append(name[1])

        self.cur.execute(
            sql, (
                tuple(condition_values),
                date_start_year, 
                date_end_year,
                tuple(condition_values), 
                date_start_quarter, 
                date_end_quarter
            )
        )
        
        sorted_data = self.cur.fetchall()
        latest_date = max(sorted_data, key=lambda x: x[3])

        return sorted_data, start_year, end_year, latest_date

    def create_quarter_table(self, names, sql, table_name, data_type, index_name):
        description = {"tenge": ["в тенге", "Год"], "percent": ["в %", "Год"], "thousand_tenge": ["в тыс тенге", "Год"]}
        excel_path = os.path.join(self.excel_path, table_name)
        workbook = Workbook()
        sheet = workbook.active

        sorted_data, start_year, end_year, latest_date = self.get_data_create_quarter_table(sql, names, 3)

        if latest_date[3].year != end_year:
            sorted_data, start_year, end_year, latest_date = self.get_data_create_quarter_table(sql, names, 4)
        
        for data in sorted_data:
            if data == latest_date:
                last_quarter = int(data[2][0])
                last_year = int(data[2].split()[-2])
                break
        

        data_for_excel = []
        for name_i, name in enumerate(names):
            index_data = []
            index_data.append(name[0])
            data_not_exist = True
            for data in sorted_data:
                if len(data[2].split()) == 2:
                    if data[0] == name[1]:
                        value = self.convert_number(data_type, data[1])
                        index_data.append(value)
                        data_not_exist = False
            if data_not_exist:
                index_data.append("")
            
            for year in range(start_year+1, last_year+1):
                end_quarter = 5
                if year == last_year: end_quarter = last_quarter + 1
                for quarter in range(1, end_quarter):
                    if name_i == 0:
                        description[data_type].append(f"{self.romam_to_arabic(quarter)} кв")
                    data_not_exist = True
                    for data in sorted_data:
                        if data[2] == f"{quarter} квартал {year} г." and data[0] == name[1]:
                            value = self.convert_number(data_type, data[1])
                            index_data.append(value)
                            data_not_exist = False
                    if data_not_exist:
                        index_data.append("")
            data_for_excel.append(index_data)
        
        year_list = list(range(start_year+1, last_year+1))
        start_cell = 3
        end_cell = 6
        sheet.cell(row=1, column=1, value="")
        sheet.cell(row=1, column=2, value=start_year)
        for i, year in enumerate(year_list):
            sheet.cell(row=1, column=start_cell, value=year)
            sheet.merge_cells(start_row=1, start_column=start_cell, end_row=1, end_column=end_cell)
            start_cell = end_cell+1
            if i+2 == len(year_list):
                end_cell += last_quarter
            else:
                end_cell += 4

        sheet.append(description[data_type])
        for data in data_for_excel:
            sheet.append(data)
                
        workbook.save(excel_path)
        workbook.close()

        print(f"Таблица {index_name} создана")