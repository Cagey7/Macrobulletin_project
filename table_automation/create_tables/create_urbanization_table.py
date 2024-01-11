from filltable import FillTable
from datetime import datetime, timedelta
import os
from openpyxl import Workbook


class CreateUrbanizationTable(FillTable):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


    def get_data_create_urbanization_table(self, sql, names, minus_num):
        current_date = datetime.now()
        start_year = current_date.year - minus_num
        end_year = start_year + 3

        date_start_year = datetime(start_year, 1, 1)
        date_end_year = datetime(start_year + 4, 1, 1) - timedelta(seconds=1)
        date_start_month = datetime(start_year + 4, 1, 1)
        date_end_month = datetime(start_year + 5, 1, 1) - timedelta(seconds=1)


        self.cur.execute(
            sql, (
                tuple([item[1] for item in names]),
                date_start_year, date_end_year,
                tuple([item[1] for item in names]),
                date_start_month, date_end_month,
            )
        )
        sorted_data = self.cur.fetchall()


        max_date = max(sorted_data, key=lambda x: x[4])[4]
        return sorted_data, end_year, start_year, max_date

    def create_urbanization_table(self, names, months_names, sql, table_name, index_name):
        excel_path = os.path.join(self.excel_path, table_name)
        workbook = Workbook()
        sheet = workbook.active

        sorted_data, end_year, start_year, max_date = self.get_data_create_urbanization_table(sql, names, 4)
        if max_date.year != end_year+1:
            sorted_data, end_year, start_year, max_date = self.get_data_create_urbanization_table(sql, names, 5)


        for data in sorted_data:
            if data[4] == max_date:
                last_month_description = data[3]
                for month in months_names:
                    if data[3].split()[0] == month[1]:
                        shorten_month = month[0]
                        break
                break


        area_type = ["Всего","городская местность"]
        data_for_excel = []
        for name in names:
            index_data = []
            index_data.append(name[0])
            data_not_exist = True

            for year in range(start_year, end_year+1):
                temp = []
                for area in area_type:
                    data_not_exist = True
                    for data in sorted_data:
                        if data[3] == f"{year} год" and data[0] == name[1] and data[1] == area:
                            temp.append(float(data[2]))
                            data_not_exist = False
                    if data_not_exist:
                        temp.append("")
                try:
                    value = float(round((temp[1]*100)/temp[0], 2))
                except:
                    index_data.append("")
                else:
                    index_data.append(value)

            temp = []
            for area in area_type:
                data_not_exist = True
                for data in sorted_data:
                    if data[3] == last_month_description and data[0] == name[1] and data[1] == area:
                        temp.append(float(data[2]))
                        data_not_exist = False
                if data_not_exist:
                    temp.append("")
            try:
                    value = float(round((temp[1]*100)/temp[0], 2))
            except:
                index_data.append("")
            else:
                index_data.append(value)
            data_for_excel.append(index_data)


        #заголовок
        year_list = list(range(start_year, end_year+2))
        start_cell = 2
        sheet.cell(row=1, column=1, value="Регионы")
        for i, year in enumerate(year_list):
            if i+1 == len(year_list):
                sheet.cell(row=1, column=start_cell, value=(f"на 1 {shorten_month}. {str(year)} г."))
            else:
                sheet.cell(row=1, column=start_cell, value=(f"{year} г."))
                start_cell += 1
            

        for data in data_for_excel:
            sheet.append(data)


        workbook.save(excel_path)
        workbook.close()

        print(f"Таблица {index_name} создана")