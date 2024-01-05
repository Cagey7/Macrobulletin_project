import calendar
import os
from openpyxl import Workbook
from filltable import FillTable
from datetime import datetime, timedelta


class CreateMonthTable(FillTable):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def get_data_create_month_table(self, sql, names, reporting_period, minus_num):
        current_date = datetime.now()
        end_month = 12
        end_year = current_date.year - minus_num + 2

        current_date = datetime.now()
        start_month = 12
        start_year = current_date.year - minus_num
        
        start_last_day = calendar.monthrange(start_year, start_month)[1]
        end_last_day = calendar.monthrange(end_year, end_month)[1]

        self.cur.execute(
            sql, (
                reporting_period,
                tuple([item[1] for item in names]),
                datetime(start_year, start_month, start_last_day), 
                datetime(end_year, end_month, end_last_day),
                reporting_period, 
                datetime(start_year, start_month, start_last_day), 
                datetime(end_year, end_month, end_last_day)
            )
        )
        sorted_data = self.cur.fetchall()
        print(sorted_data)
        print(start_year)
        print(start_month)
        print(end_year)
        print(end_month)

        return sorted_data, start_year, start_month, end_year, end_month

    def create_month_table(self, names, months, sql, table_name, reporting_period, index_name):
        excel_path = os.path.join(self.excel_path, table_name)

        workbook = Workbook()
        sheet = workbook.active


        sorted_data, start_year, start_month, end_year, end_month = self.get_data_create_month_table(sql, names, reporting_period, 2)
        dates = [item[3] for item in sorted_data]
        latest_month = max(dates).month
        latest_year = max(dates).year

        if latest_month != start_year:
            sorted_data, start_year, start_month, end_year, end_month = self.get_data_create_month_table(sql, names, reporting_period, 3)

        if latest_month < end_month:
            end_month = latest_month
        if latest_year < end_year:
            end_year = latest_year

        
        months_list = [""]

        data_for_excel = []
        for bull_name_i, name in enumerate(names):
            index_data = []
            index_data.append(name[0])
            for year in range(start_year, end_year+1):
                for i, m in enumerate(months, start=1):
                    if year == start_year and i < start_month: continue
                    if year == end_year and i > end_month: continue
                    # получешние месяцев для заголовка
                    if bull_name_i == 0:
                        months_list.append(m[0])
                    
                    data_not_exist = True
                    for data in sorted_data:
                        t_month, t_year, _ = data[2].split()
                        t_year = int(t_year)
                        if name[1] == data[0] and t_month == m[1] and t_year == year:
                            index_data.append(float(data[1]))
                            data_not_exist = False
                    if data_not_exist:
                        index_data.append("")
            data_for_excel.append(index_data)

        # заголовок
        year_list = list(range(start_year, end_year + 1))
        start_cell = 2
        end_cell = 12-start_month+start_cell
        sheet.cell(row=1, column=1, value="Показатели")
        for i, year in enumerate(year_list):
            sheet.cell(row=1, column=start_cell, value=year)
            if start_year == end_year and start_month == end_month: break
            sheet.merge_cells(start_row=1, start_column=start_cell, end_row=1, end_column=end_cell)
            start_cell = end_cell+1
            if i+2 == len(year_list):
                end_cell += end_month
            else:
                end_cell += 12


        sheet.append(months_list)
        for data in data_for_excel:
            sheet.append(data)
        workbook.save(excel_path)
        workbook.close()

        print(f"Таблица {index_name} создана")
