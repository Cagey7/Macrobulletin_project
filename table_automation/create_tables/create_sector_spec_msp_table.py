from filltable import FillTable
from datetime import datetime, timedelta
import os
from openpyxl import Workbook


class CreateMSPTable(FillTable):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


    def get_data_create_sector_spec_msp_table(self, sql, query_names, minus_num):
        current_date = datetime.now()
        start_year = current_date.year - minus_num
        end_year = start_year + 1

        date_start_year = datetime(start_year, 1, 1)
        date_end_year = datetime(start_year + 3, 1, 1) - timedelta(seconds=1)


        self.cur.execute(
            sql, (
                tuple(query_names[16:20]),
                date_start_year, date_end_year,
                tuple(query_names[20:]),
                date_start_year, date_end_year,
                tuple(query_names[:16]),
                date_start_year, date_end_year,
            )
        )
        sorted_data = self.cur.fetchall()

        max_date = max(sorted_data, key=lambda x: x[4])[4]
        return sorted_data, end_year, start_year, max_date

    def create_sector_spec_msp_table(self, names, query_names, sql, table_name, index_name):
        excel_path = os.path.join(self.excel_path, table_name)
        workbook = Workbook()
        sheet = workbook.active

        sorted_data, end_year, start_year, max_date = self.get_data_create_sector_spec_msp_table(sql, query_names, 2)
        if max_date.year != end_year:
            sorted_data, end_year, start_year, max_date = self.get_data_create_sector_spec_msp_table(sql, query_names, 3)

        gen_msp = []
        enterprise_dimension = ["Малые ", "Средние "]

        for ent_type in enterprise_dimension:
            for year in range(start_year, end_year+1):
                for data in sorted_data:
                    if data[3] == f"{year} год" and data[0] == "Всего" and data[1] == ent_type:
                        gen_msp.append(data[2])
                        break

        data_for_excel = []

        for name in names:
            index_data = []
            index_data.append(name[0])
            data_not_exist = True
            indeces = [0, 1, 1, 2]
            j = 0
            for i, ent_type in enumerate(enterprise_dimension):
                for year in range(start_year, end_year+1):
                    data_not_exist = True
                    for data in sorted_data:
                        if data[3] == f"{year} год" and data[0] == name[1] and data[1] == ent_type:
                            value = round(float(data[2])/float(gen_msp[i+indeces[j]]) * 100, 1)
                            index_data.append(value)
                            data_not_exist = False
                    if data_not_exist:
                        index_data.append("")
                    j += 1

            
            data_for_excel.append(index_data)


        #заголовок
        year_list = list(range(start_year, end_year+1))
        start_cell = 2
        end_cell = 3
        sheet.cell(row=1, column=1, value="в %")
        sheet.cell(row=1, column=2, value="субъекты малого предпринимательства")
        sheet.cell(row=1, column=4, value="субъекты среднего предпринимательства")
        sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
        sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
        for _ in range(2):
            for  year in year_list:
                    sheet.cell(row=2, column=start_cell, value=(f"{year} г."))
                    start_cell += 1
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            

        for data in data_for_excel:
            sheet.append(data)


        workbook.save(excel_path)
        workbook.close()

        print(f"Таблица {index_name} создана")