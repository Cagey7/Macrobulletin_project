from filltable import FillTable
from datetime import datetime, timedelta
import os
from openpyxl import Workbook


class CreatePopBegin2Table(FillTable):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


    def get_data_create_pop_begin_2_table(self, sql, names, minus_num):
        current_date = datetime.now()
        year = current_date.year - minus_num

        date_start_year = datetime(year, 1, 1)
        date_end_year = datetime(year + 1, 1, 1) - timedelta(seconds=1)


        self.cur.execute(
            sql, (
                tuple([item[1] for item in names]),
                date_start_year, date_end_year,
            )
        )
        sorted_data = self.cur.fetchall()

        try:
            max_date = max(sorted_data, key=lambda x: x[5])[5]
        except:
            max_date = datetime(1, 1, 1)
        return sorted_data, year, max_date

    def create_pop_begin_2_table(self, names, sql, table_name, index_name):
        excel_path = os.path.join(self.excel_path, table_name)
        workbook = Workbook()
        sheet = workbook.active

        sorted_data, year, max_date = self.get_data_create_pop_begin_2_table(sql, names, 0)
        if max_date.year != year+1:
            sorted_data, year, max_date = self.get_data_create_pop_begin_2_table(sql, names, 1)


        population_group = ["Все группы","Моложе трудоспособного населения","Трудоспособное население","Старше трудоспособного населения"]
        genders = ["Всего","Мужской","Женский"]
        data_for_excel = []
        for name in names:
            index_data = []
            index_data.append(name[0])
            data_not_exist = True


            for group in population_group:
                for gender in genders:
                    data_not_exist = True
                    for data in sorted_data:
                        if data[0] == name[1] and data[2] == group and data[1] == gender:
                            value = float(data[3])
                            index_data.append(value)
                            data_not_exist = False
                    if data_not_exist:
                            index_data.append("")

            
            data_for_excel.append(index_data)

        description1 = ["Регионы","Всего","","","в том числе в возрасте"]
        description2 = ["","","","","0-15","","","16-62(59)","","","63(60)+*"]
        description3 = [""] + ["всего", "мужчины", "женщины"] * 4

        #заголовок
        for i, item in enumerate(description1):
            sheet.cell(row=1, column=i+1, value=item)
        for i, item in enumerate(description2):
            sheet.cell(row=2, column=i+1, value=item)
        for i, item in enumerate(description3):
            sheet.cell(row=3, column=i+1, value=item)
        sheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
        sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=4)
        sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=13)
        sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
        sheet.merge_cells(start_row=2, start_column=8, end_row=2, end_column=10)
        sheet.merge_cells(start_row=2, start_column=11, end_row=2, end_column=13)
        
 
        for data in data_for_excel:
            sheet.append(data)


        workbook.save(excel_path)
        workbook.close()

        print(f"Таблица {index_name} создана")