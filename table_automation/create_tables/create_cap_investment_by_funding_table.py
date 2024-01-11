from filltable import FillTable
from datetime import datetime, timedelta
import os
from openpyxl import Workbook


class CreateCapInvestmentByFundingTable(FillTable):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


    def get_data_create_cap_investment_by_funding_table(self, sql, names, accum_type, minus_num):
        current_date = datetime.now()
        start_year = current_date.year - minus_num
        end_year = start_year + 3


        date_start_quarter = datetime(start_year, 1, 1)
        date_end_quarter= datetime(start_year + 4, 1, 1) - timedelta(seconds=1)


        self.cur.execute(
            sql, (
                tuple([item[1] for item in names]),
                date_start_quarter, date_end_quarter,
                date_start_quarter, date_end_quarter,
                date_start_quarter, date_end_quarter,
            )
        )
        
        sorted_data = self.cur.fetchall()
        max_date, accum_description = self.last_quarter_month(accum_type, sorted_data, date_end_quarter, 3)

        return sorted_data, end_year, start_year, max_date, accum_description

    def create_cap_investment_by_funding_table(self, names, accum_type, sql, table_name, data_type, index_name):
        excel_path = os.path.join(self.excel_path, table_name)
        workbook = Workbook()
        sheet = workbook.active

        sorted_data, end_year, start_year, max_date, accum_description = self.get_data_create_cap_investment_by_funding_table(sql, names, accum_type, 3)
        if max_date.year != end_year+1:
            sorted_data, end_year, start_year, max_date, accum_description = self.get_data_create_cap_investment_by_funding_table(sql, names, accum_type, 4)


        if data_type == "million_tenge":
            data_type_written = "Млн. тенге"
        else:
            data_type_written = ""

        if accum_type == "quarter":
            desc_ending = "кв."
        elif accum_type == "month":
            desc_ending = "мес."
        else:
            desc_ending = ""
        
        for data in sorted_data:
            if data[3] == max_date:
                accum_quarter = data[2]
                break
        
        percent_sum = []
        for year in range(start_year, end_year + 1):
            for data in sorted_data:
                if data[0] == "Сумма" and (data[2] == f"Январь - Декабрь {year} г. ({desc_ending})" or data[2] == accum_quarter):
                    percent_sum.append(data[1])
                    break

        data_for_excel = []
        for name in names:
            index_data = []
            index_data.append(name[0])
            data_not_exist = True

            for i, year in enumerate(range(start_year, end_year)):
                data_not_exist = True
                for data in sorted_data:
                    if data[2] == f"Январь - Декабрь {year} г. ({desc_ending})" and data[0] == name[1]:
                        value = self.convert_number(data_type, data[1])
                        index_data.append(value)
                        index_data.append(round((data[1]/percent_sum[i]) * 100, 1))
                        data_not_exist = False
                if data_not_exist:
                    index_data.append("")


            data_not_exist = True
            for data in sorted_data:
                if data[2] == accum_quarter and data[0] == name[1]:
                    value = self.convert_number(data_type, data[1])
                    index_data.append(value)
                    index_data.append(round((data[1]/percent_sum[3]) * 100, 1))
                    data_not_exist = False
            if data_not_exist:
                index_data.append("")
            
            data_for_excel.append(index_data)

        description = [data_type_written] + ["", "доля (%)"] * 3 + [accum_description, "г/г (%)"]

        #заголовок
        year_list = list(range(start_year, end_year+1))
        start_cell = 2
        end_cell = 3
        sheet.cell(row=1, column=1, value="Показатели")
        for year in year_list:
            sheet.cell(row=1, column=start_cell, value=(f"{year} г."))
            sheet.merge_cells(start_row=1, start_column=start_cell, end_row=1, end_column=end_cell)
            start_cell = end_cell+1
            end_cell += 2
            

        sheet.append(description)
        for data in data_for_excel:
            sheet.append(data)


        workbook.save(excel_path)
        workbook.close()

        print(f"Таблица {index_name} создана")